from seaborn import scatterplot, lineplot
import pandas as pd
import numpy as np
from lmfit import Model, Parameters
import matplotlib.ticker as mtick
import matplotlib.pyplot as plt
import openpyxl

def FitSpit(dat, parameters, dated):
    def func(x, u, s, l, i):
        return ((l-u)/(1.0+((x/i)**s))) + u
    init_model = Model(func)

    df = dat

    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    worksheet['A1'] = 'sample ID'
    worksheet['B1'] = 'chi-squared'
    worksheet['C1'] = 'r-squared'
    worksheet['D1'] = 'upper limit'
    worksheet['E1'] = 'slope'
    worksheet['F1'] = 'lower limit'
    worksheet['G1'] = 'ND50'
    worksheet['H1'] = 'ND80'
    worksheet['I1'] = 'plots'
    
    # sampleids = []
    # chis = []
    # rs = []
    # imgs = []
    # variables = []
    # nd80 = []
    sampletrack = 0
        
    for index, plate in enumerate(df['plateID'].unique()):
        print("moving on to " + str(plate))
        df_filtered = df[df['plateID'] == plate]

        # calculate average foci in VOC to normalize foci reduction
        VOC_avg = df_filtered[df_filtered['type']=='VOC'].foci_num.mean()
        neg_avg = df_filtered[df_filtered['type']=='negative'].foci_num.mean()

        df_filtered = df_filtered[(df_filtered['type']!='negative') & (df_filtered['type']!='VOC')].copy()
        
        df_filtered['foci_red']= 1 - (df_filtered['foci_num']/VOC_avg).copy()

        numofsamples = len(df['type'].unique())
        
        for num in np.arange(1, max(df_filtered['sample_num'])+1, 1):
            sample = df_filtered[df_filtered['sample_num']==num]
            samplename = [str(sampleid) for sampleid in sample['type'].unique()][0]
            xpoints = [fold_dil for fold_dil in sample['fold_dil']]
            xpoints = np.arange(min(xpoints)-10, max(xpoints)+4000, max(xpoints)/1000)

            ydat = init_model.fit(
                    sample['foci_red'],
                    params = parameters,
                    x=sample['fold_dil'], 
                    method = 'nelder'
                )

            sampletrack += 1

        # set up figure, axes
            plt.figure(sampletrack, dpi=70)
            plt.figure(sampletrack).set_figwidth(6)
            plt.figure(sampletrack).set_figheight(5.2)
          
            axie = plt.axes()
            axie.set_ylabel('Foci reduction', fontsize=12)
            axie.set_xlabel('Fold dilution', fontsize=12)
            axie.set_xscale('log')
            axie.tick_params(axis = 'x', which='major', length=12, width=2)
            axie.tick_params(axis = 'x', which='minor', length= 5, width=2)
            axie.set_yticks(np.arange(-0.2, 1.1, 0.1))
            axie.set_ylim([-0.2, 1.1])
            # axie.set_xlim([10**1, 13**4])
            axie.set_xlim([min(xpoints)-5, max(xpoints)+4500])
            axie.set_title(samplename, fontsize=15)
            axie.yaxis.set_major_formatter(mtick.PercentFormatter(1))
                           
            scatterplot(ax=axie,
                        x=sample['fold_dil'],
                        y=sample['foci_red'],
                        color='blue',
                        alpha=0.6
                        )

            lineplot(ax=axie, 
                     y=ydat.eval(x=xpoints),
                     x=xpoints,
                     color='black'
                     )

            plt.tight_layout()
  
            plt.savefig(str(sampletrack) + samplename + '.png')

            worksheet.cell(row=sampletrack + 1, column=1).value = samplename

            worksheet.add_image(openpyxl.drawing.image.Image(str(sampletrack) + samplename + '.png'), anchor='I'+str(sampletrack+1))

            worksheet.cell(row=sampletrack + 1, column=2).value = ydat.chisqr

            worksheet.cell(row=sampletrack + 1, column=3).value = ydat.rsquared

            worksheet.cell(row=sampletrack + 1, column=8).value = (((80)/(20))**(1/ydat.best_values['s'])) * (ydat.best_values['i'])

            # imgs += [samplename + '.png']
        #### KEEEP THIS !!!!!!!!!!!!!!
            for x, dic in enumerate(ydat.best_values.values()): 
                worksheet.cell(row=sampletrack+1, column=4+x).value = dic

            # nd80.append(
            #         (((80)/(20))**(1/ydat.best_values['s'])) * (ydat.best_values['i']))

            # chis.append(ydat.chisqr)
        
            # rs.append(ydat.rsquared)

            # sampleids.append(samplename)

    print(sampletrack)

    for row in range(2, sampletrack+2):
        worksheet.row_dimensions[row].height = 280
        worksheet.column_dimensions['I'].width = 70

    # for index, sampleid in enumerate(sampleids):
    #     worksheet.cell(row=index+2, column=1).value = sampleid
                
    # for index, image in enumerate(imgs):
    #     worksheet.add_image(openpyxl.drawing.image.Image(image), anchor='I'+str(index+2))
                       
    # for i, chi in enumerate(chis):
    #     worksheet.cell(row=num+4*index+2, column=2).value = chi
        
    # for i, rsquared in enumerate(rs):
    #     worksheet.cell(row=num*4*index+2, column=3).value = rsquared

    # for i, nd80 in enumerate(nd80):
    #     worksheet.cell(num+4*index+2, column = 8).value = nd80

        
    workbook.save(str(dated) +'.xlsx')
    plt.close()


