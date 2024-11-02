from seaborn import scatterplot, lineplot
import numpy as np
import matplotlib.ticker as mtick
import matplotlib.pyplot as plt
from lmfit import Model, Parameters
import openpyxl
import argparse


def define_fit_params():
    """
    NOTE: these initial guesses for parameters adhere to the
    analysis performed by the Greninger Lab for the Cascadia study.
    They are hardcoded to maintain compliance.
    """

    params = Parameters()
    params.add("u", value=1, min=0.95, max=1.05)
    params.add("s", value=0.8, min=-1.2, max=0.7)
    params.add("l", value=0, min=-0.05, max=0.05)
    params.add("i", value=50, min=0, max=np.inf)

    return params


PARAMETERS = define_fit_params()


def curve_function(x, u, s, l, i):
    """Define the function for curve fitting."""
    return ((l - u) / (1.0 + ((x / i) ** s))) + u


def setup_worksheet(worksheet):
    """Set up the header row in the worksheet."""
    headers = [
        "sample ID",
        "chi-squared",
        "r-squared",
        "upper limit",
        "slope",
        "lower limit",
        "ND50",
        "ND80",
        "plots",
    ]
    for idx, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=idx, value=header)


def init_empty_plot(sampletrack, samplename, xpoints):
    """Set up the plot figure and axes."""
    plt.figure(sampletrack, dpi=70)
    plt.figure(sampletrack).set_figwidth(6)
    plt.figure(sampletrack).set_figheight(5.2)

    axie = plt.axes()
    axie.set_ylabel("Foci reduction", fontsize=12)
    axie.set_xlabel("Fold dilution", fontsize=12)
    axie.set_xscale("log")
    axie.tick_params(axis="x", which="major", length=12, width=2)
    axie.tick_params(axis="x", which="minor", length=5, width=2)
    axie.set_yticks(np.arange(-0.2, 1.1, 0.1))
    axie.set_ylim([-0.2, 1.1])
    axie.set_xlim([min(xpoints) - 5, max(xpoints) + 4500])
    axie.set_title(samplename, fontsize=15)
    axie.yaxis.set_major_formatter(mtick.PercentFormatter(1))
    return axie


def fill_worksheet(worksheet, sampletrack, samplename, ydat):
    """Fill the worksheet with data for a sample."""
    worksheet.cell(row=sampletrack + 1, column=1).value = samplename
    worksheet.add_image(
        openpyxl.drawing.image.Image(f"{sampletrack}{samplename}.png"),
        anchor=f"I{sampletrack+1}",
    )
    worksheet.cell(row=sampletrack + 1, column=2).value = ydat.chisqr
    worksheet.cell(row=sampletrack + 1, column=3).value = ydat.rsquared
    worksheet.cell(row=sampletrack + 1, column=8).value = (
        ((80) / (20)) ** (1 / ydat.best_values["s"])
    ) * (ydat.best_values["i"])

    for x, value in enumerate(ydat.best_values.values()):
        worksheet.cell(row=sampletrack + 1, column=4 + x).value = value


def fit_sample(sample, parameters, sampletrack):
    """Process a single sample and create its plot."""
    samplename = str(sample["type"].unique()[0])
    xpoints = sample["fold_dil"].tolist()
    xpoints = np.arange(min(xpoints) - 10, max(xpoints) + 4000, max(xpoints) / 1000)

    init_model = Model(curve_function)
    ydat = init_model.fit(
        sample["foci_red"],
        params=parameters,
        x=sample["fold_dil"],
        method="nelder",  ## calculate reduction in foci relative to dilution
    )

    axie = init_empty_plot(sampletrack, samplename, xpoints)

    scatterplot(
        ax=axie, x=sample["fold_dil"], y=sample["foci_red"], color="blue", alpha=0.6
    )
    lineplot(ax=axie, y=ydat.eval(x=xpoints), x=xpoints, color="black")

    plt.tight_layout()
    plt.savefig(
        f"{sampletrack}{samplename}.png"
    )  # need to save image for embedding into excel

    return ydat, samplename


def FitSpit(dat, output_name):
    df = dat
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    setup_worksheet(worksheet)

    sampletrack = 0

    for plate in df["plateID"].unique():
        print(f"Processing plate: {plate}")
        df_filtered = df[df["plateID"] == plate]

        VOC_avg = df_filtered[df_filtered["type"] == "VOC"].foci_num.mean()
        df_filtered = df_filtered[
            (df_filtered["type"] != "negative") & (df_filtered["type"] != "VOC")
        ].copy()
        df_filtered["foci_red"] = 1 - (df_filtered["foci_num"] / VOC_avg)

        for num in range(1, df_filtered["sample_num"].max() + 1):
            sample = df_filtered[df_filtered["sample_num"] == num]
            sampletrack += 1

            ydat, samplename = fit_sample(sample, PARAMETERS, sampletrack)
            fill_worksheet(worksheet, sampletrack, samplename, ydat)

    print(f"Total samples processed: {sampletrack}")

    # resize cells containing plot images to avoid overflow
    for row in range(2, sampletrack + 2):
        worksheet.row_dimensions[row].height = 280
    worksheet.column_dimensions["I"].width = 70

    workbook.save(f"{output_name}.xlsx")
    plt.close("all")


def parse_arguments():
    parser = argparse.ArgumentParser(
        description="Process and analyze foci reduction data."
    )
    parser.add_argument(
        "in", help="Path of the .xlsx sum file generated from reconf.py"
    )
    parser.add_argument("out", help="Filename/path for the output file with curve fits")
    return parser.parse_args()


def main():
    args = parse_arguments()
    FitSpit(args.input_file, args.output_file)


if __name__ == "__main__":
    main()
