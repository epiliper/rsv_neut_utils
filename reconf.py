#!/usr/bin/env python3

"""
Author: Eli Piliper

This script converts .xlsx files of plate counts from the CTL S6 imager
into per-sample reports containing sample ID, dilution, spot counts, and plateID.

The output of this script should be used with the Fit.py script to generate 4-parameter logistic
curves for obtaining ND50 and ND80 values from data generated with the Greninger Lab's
RSV focus-reduction neutralization assay (RSV FRNT).

Note that the CTL S6 imager outputs data in .xls format. These files will need to be
converted to .xlsx prior to running this and other scripts.
"""

import openpyxl as xl
from typing import List, Tuple
import logging
import argparse
import sys
import os

# Set up logging
logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)

# Constants
ROWS = ["A", "B", "C", "D", "E", "F", "G", "H"]
COLS = list(range(1, 13))


def load_workbooks(file_paths: List[str]) -> List[xl.Workbook]:
    """Load Excel workbooks from given file paths."""
    try:
        return [xl.load_workbook(file) for file in file_paths]
    except Exception as e:
        logging.error(f"Error loading workbooks: {e}")
        raise


def get_sheet_data(sheets: List[xl.Worksheet]) -> List[Tuple]:
    """Extract data from specific cells in the sheets."""
    return [
        sheet.iter_cols(min_col=3, max_col=14, min_row=55, max_row=62)
        for sheet in sheets
    ]


def generate_labels(num_plates: int) -> List[str]:
    """Generate well labels for all plates."""
    return [f"{row}{col}" for _ in range(num_plates) for col in COLS for row in ROWS]


def create_new_workbook() -> Tuple[xl.Workbook, xl.Worksheet]:
    """Create a new workbook and set up the header row."""
    workbook = xl.Workbook()
    sheet = workbook.active
    headers = ["foci_num", "fold_dil", "type", "wellID", "sample_num", "plateID"]
    for col, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col, value=header)
    return workbook, sheet


def process_sample_data(sample_sheets: List[xl.Worksheet]) -> List[str]:
    """Process sample data from the sample sheets."""
    samples = []
    for sheet in sample_sheets:
        column_c = sheet["C"]
        for cell in column_c[1:]:  # Start from the second row (index 1)
            if cell.value is not None:
                samples.append(cell.value)
            else:
                break  # Stop when we hit empty cell (end of run)
    return samples * 2  # samples were run in duplicate, so make two sets of metadata


def generate_dilutions(num_samples: int) -> List[int]:
    """
    Generate dilution values for samples
    TODO: read runsheets for dilution information to avoid hardcoding
    """
    dilutions = [20, 60, 180, 540, 1620, 4860]
    return dilutions * (num_samples // 6)


def identify_special_wells(num_plates: int) -> Tuple[List[int], List[int]]:
    """
    Identify wells for negatives and VOCs.

        VOCs = rows B-G of columns 10 and 11
        Negatives = rows A and H, and columns 1 and 12

    """
    total_wells = 96 * num_plates
    negatives = [
        i
        for i in range(1, total_wells + 1)
        if i > 8 and ((i - 1) % 8 == 0 or (i - 2) % 8 == 0)
    ]
    negatives.extend([a + 96 * b for a in range(1, 9) for b in range(num_plates)])
    negatives.extend([a + 96 * b for a in range(88, 97) for b in range(num_plates)])
    negatives.sort()

    vocs = [a + 96 * b for a in range(73, 79) for b in range(num_plates)]
    vocs.extend([a + 96 * b for a in range(81, 87) for b in range(num_plates)])
    vocs.sort()

    return negatives, vocs


def ctlimport(workbooks: List[str], ctl: List[str], export: str):
    """Main function to process CTL import and generate the output file."""
    try:
        # load ctl data
        ctl_books = load_workbooks(ctl)
        ctl_sheets = [book.active for book in ctl_books]
        sample_sheets = [xl.load_workbook(wb)["Serum Dilution"] for wb in workbooks]

        num_plates = len(ctl_sheets)
        logging.info(f"Processing {num_plates} plates")

        # generate labels and plate IDs
        labels = generate_labels(num_plates)
        plate_ids = [
            os.path.basename(countsheet).replace(".xlsx", "") for countsheet in ctl
        ]

        # create new workbook
        new_workbook, new_sheet = create_new_workbook()

        # process sample data
        samples = process_sample_data(sample_sheets, num_plates)
        dilutions = generate_dilutions(len(samples))

        negatives, vocs = identify_special_wells(num_plates)

        # fill ouput with sample, dilution, neutralizing activity
        sheet_data = get_sheet_data(ctl_sheets)
        for index, sheet in enumerate(sheet_data):
            for i, old_col in enumerate(sheet, 0):
                for row, count in enumerate(old_col):
                    new_sheet.cell(
                        row=(row + (8 * i)) + 1 + (96 * index),
                        column=1,
                        value=count.value,
                    )

        total_wells = 96 * num_plates
        sample_index = 0
        dilution_index = 0

        for well in range(1, total_wells + 1):
            row = new_sheet[well + 1]  # +1 because row 1 is the header

            if well in negatives:
                row[2].value = "negative"  # type
                row[1].value = None  # fold_dil
            elif well in vocs:
                row[2].value = "VOC"
                row[1].value = None
            else:
                row[2].value = samples[sample_index]
                row[1].value = dilutions[dilution_index]
                sample_index = (sample_index + 1) % len(samples)
                dilution_index = (dilution_index + 1) % len(dilutions)

            row[3].value = labels[well - 1]  # well number on run plate
            row[4].value = (well - 1) // 6 + 1  # sample number on run plate
            row[5].value = plate_ids[(well - 1) // 96]  # plate ID

        # write output
        new_workbook.save(f"{export}.xlsx")
        logging.info(f"Run saved to {export}.xlsx")

    except Exception as e:
        logging.error(f"Run import and analysis failed!: {e}")
        raise


def parse_arguments():
    parser = argparse.ArgumentParser(
        description="Process CTL import and generate output file."
    )
    parser.add_argument(
        "-w", "--workbooks", nargs="+", required=True, help="List of workbook files"
    )
    parser.add_argument(
        "-c", "--ctl", nargs="+", required=True, help="List of CTL xlsx files"
    )
    parser.add_argument(
        "-e",
        "--export",
        required=True,
        help="Name of the export file",
    )
    return parser.parse_args()


def main():
    """Main entry point of the script."""
    args = parse_arguments()
    try:
        ctlimport(args.workbooks, args.ctl, args.export)
    except Exception as e:
        logging.error(f"Run analysis failed! {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
